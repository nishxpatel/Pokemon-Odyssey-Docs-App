#!/usr/bin/env python3
"""
ETL: parse Pokémon Odyssey v4.1.1 spreadsheets into JSON for the site.

Outputs (written to ./site/data/):
  pokedex.json    — array of species objects: dex, name, slug, types, abilities,
                    stats (odyssey + vanilla), moves, locations, evolution chain,
                    evolution_items, family, variant_sprite, is_variant, etc.
  moves.json      — {"moves": [...]} — all moves (custom Odyssey + baseline from
                    PokeAPI) each with name, slug, type, category, power,
                    accuracy, pp, effect, kind, is_custom, used_by[]
  abilities.json  — {"abilities": [...]} — same shape as moves minus combat stats
  items.json      — {"items": [...], "tutors": [...]} — items keyed by source
                    kind (location / shop / pickup / gather / tm), plus the move
                    tutor list
  meta.json       — game version, full type list (19 types inc. Aether),
                    19×19 type_chart (non-1× interactions only), summary counts

Also extracts embedded variant sprites from the Etrian Variants sheet into
./site/assets/variants_original/<slug>.png (raw), then strips backgrounds into
./site/assets/variants/<slug>.png (served by the site).

Run with: python3 scripts/build_data.py  (from the project root)
"""

import json
import re
import sys
import unicodedata
import xml.etree.ElementTree as ET
import zipfile
from datetime import datetime
from pathlib import Path

import openpyxl

HERE = Path(__file__).parent
ROOT = HERE.parent
DOCS_DIR = ROOT / "docs"
STATS_XLSX  = DOCS_DIR / "Pokémon Stats, Learnset etc (v4.1.1).xlsx"
WILD_XLSX   = DOCS_DIR / "Wild encounters, Items and TMs (v4.1.1).xlsx"
BOSSES_XLSX = DOCS_DIR / "Level Cap, Boss, Miniboss, Sea Map, Sidequests (v4.1.1).xlsx"
SITE_DIR    = ROOT / "site"
OUT_DIR     = SITE_DIR / "data"
VARIANTS_DIR = SITE_DIR / "assets" / "variants"
ORIGINALS_DIR = SITE_DIR / "assets" / "variants_original"

STAT_SHEETS = ["#1-151", "#152-251", "#252-386", "4th Gen", "Paradox"]

LABELS = {"TYPE:", "ABILITY:", "EVOLUTION:", "MOVES"}
STAT_HEADERS = {"HP", "ATK", "DEF", "SP.ATK", "SP.DEF", "SPD", "TOTAL", "GAME"}

# Branching evolutions that the spreadsheet layout cannot express unambiguously:
# the source species shows a combined "A/B" evolution condition in one cell, with
# only one branch target placed in the same band (col G). The second branch target
# lives in a separate sheet as an apparent standalone. Each entry here overrides the
# combined condition on the existing edge AND adds the missing second branch edge.
#
# Also includes within-sheet branching where the layout isolates the source species
# in its own block (e.g. Wurmple is alone at row 132 while its two targets each
# appear on separate rows further down).
#
# Format: from_key -> [(to_key, condition_for_that_branch), ...]
EXPLICIT_BRANCHES = {
    "GLOOM":    [("VILEPLUME",  "Leaf Stone"), ("BELLOSSOM", "Sun Stone")],
    "POLIWHIRL":[("POLIWRATH",  "Water Stone"), ("POLITOED",  "Link Stone")],
    "SLOWPOKE": [("SLOWBRO",   "LV.37"),       ("SLOWKING",  "Link Stone")],
    "SCYTHER":  [("KLEAVOR",   "B. Augurite"), ("SCIZOR",    "Link Stone")],
    "KIRLIA":   [("GARDEVOIR", "LV.30"),       ("GALLADE",   "Dawn Stone")],
    "SNORUNT":  [("GLALIE",    "LV.42"),       ("FROSLASS",  "Dawn Stone")],
    # Wurmple: branching in-sheet but isolated in its own block; Silcoon/Cascoon
    # appear on separate rows so no band spans Wurmple to either target.
    "WURMPLE":  [("SILCOON",   "LV. 7"),       ("CASCOON",   "LV. 7")],
    # Nincada: the band detector produces stages=[NINCADA, NINJASK, SHEDINJA] and
    # tries NINJASK→SHEDINJA as the second sequential edge. But NINJASK has no
    # evolves_at so that edge is dropped and Shedinja ends up standalone. The
    # correct structure is both NINJASK and SHEDINJA branching from NINCADA at
    # the same level-up event (Shedinja requires an empty party slot + spare Ball).
    "NINCADA":  [("NINJASK",   "LV. 20"),      ("SHEDINJA",  "LV. 20 (spare slot)")],
}

# Cross-sheet/cross-generation linear evolutions.
# The band parser only connects species that appear on the same spreadsheet row.
# Any evolution where the source and target live on different sheets (e.g. a Gen 1
# Pokémon evolving into a Gen 4 form) will have no band, so no edge is inferred.
# These entries wire them explicitly.  Conditions are taken verbatim from the
# source species' evolves_at cell in the workbook.
#
# Format: from_key -> [(to_key, condition), ...]
CROSS_SHEET_EVOS = {
    # Gen 1 base → Gen 2 evolution  (#1-151 → #152-251)
    "CHANSEY":    [("BLISSEY",    "Happiness")],
    "GOLBAT":     [("CROBAT",     "Happiness")],
    "ONIX":       [("STEELIX",    "Link Stone")],
    "SEADRA":     [("KINGDRA",    "Link Stone")],
    "PORYGON":    [("PORYGON2",   "Upgrade")],
    # Gen 2 baby → Gen 1 mid-stage  (#152-251 → #1-151)
    "SMOOCHUM":   [("JYNX",       "LV.30")],
    "ELEKID":     [("ELECTABUZZ", "LV.30")],
    "MAGBY":      [("MAGMAR",     "LV.30")],
    # Gen 1 base/mid → Gen 4 final  (#1-151 → 4th Gen)
    "TANGELA":    [("TANGROWTH",  "LV.38")],
    "LICKITUNG":  [("LICKILICKY", "LV.33")],
    "RHYDON":     [("RHYPERIOR",  "Protector")],
    "ELECTABUZZ": [("ELECTIVIRE", "Electirizer")],
    "MAGMAR":     [("MAGMORTAR",  "Magmarizer")],
    "MAGNETON":   [("MAGNEZONE",  "Thunderstone")],
    # Gen 2 base → Gen 4 evolution  (#152-251 → 4th Gen)
    "SNEASEL":    [("WEAVILE",    "LV.35")],
    "MISDREAVUS": [("MISMAGIUS",  "Dusk Stone")],
    "AIPOM":      [("AMBIPOM",    "LV.30")],
    "YANMA":      [("YANMEGA",    "LV.33")],
    "MURKROW":    [("HONCHKROW",  "Dusk Stone")],
    "GLIGAR":     [("GLISCOR",    "LV.35")],
    "PILOSWINE":  [("MAMOSWINE",  "LV.45")],
    "PORYGON2":   [("PORYGONZ",   "Dubious Disc")],
    "TOGETIC":    [("TOGEKISS",   "Shiny Stone")],
    # Gen 3 base → Gen 4 evolution  (#252-386 → 4th Gen)
    "DUSCLOPS":   [("DUSKNOIR",   "Reaper Cloth")],
    "NOSEPASS":   [("PROBOPASS",  "Thunderstone")],
    "ROSELIA":    [("ROSERADE",   "Shiny Stone")],
}

# Known data errors in the workbook: evolves_at values that are wrong due to
# copy-paste mistakes in the source spreadsheet. Applied to all_species before
# the evolution graph is built so both the graph and the JSON output are clean.
# Format: canon_key -> corrected evolves_at value (None = terminal/no evolution)
EVOLVES_AT_OVERRIDES = {
    # Armaldo #348: spreadsheet duplicates Anorith's 'LV.40' into Armaldo's
    # adjacent evolution cell. Armaldo is a final form with no evolution.
    "ARMALDO": None,
}

# Items that can trigger evolution. The UI links these to item pages.
EVOLUTION_ITEMS = {
    "Fire Stone", "Water Stone", "Leaf Stone", "Thunder Stone", "Thunderstone",
    "Moon Stone", "Sun Stone", "Dawn Stone", "Dusk Stone", "Shiny Stone",
    "Ice Stone", "Oval Stone",
    "Link Stone",
    "Protector", "Dubious Disc", "Electirizer", "Magmarizer", "Reaper Cloth",
    "Prism Scale", "Upgrade",
    "B. Augurite", "Black Augurite",
}

# -----------------------------------------------------------------------------
# Name utilities
# -----------------------------------------------------------------------------

def slugify(name: str) -> str:
    s = (name or "").replace("⭐", "").strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = s.lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s

def sprite_slug(name: str) -> str:
    """Pokémon Showdown sprite slug. For variants, falls back to the base species."""
    n = (name or "").replace("⭐", "").strip()
    if "Nidoran" in n and "(F)" in n: n = "Nidoran F"
    elif "Nidoran" in n and "(M)" in n: n = "Nidoran M"
    else: n = re.sub(r"\s*\(.*?\)", "", n)
    n = n.replace("♀", "f").replace("♂", "m")
    n = n.lower()
    return re.sub(r"[^a-z0-9]+", "", n)

def canon(name) -> str:
    s = unicodedata.normalize("NFKD", str(name or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]+", "", s.upper())

def is_species_header(v) -> bool:
    if not (v and isinstance(v, str)): return False
    u = v.strip()
    if not u or u.upper() != u: return False
    if u in LABELS or u in STAT_HEADERS: return False
    if u in {"EVOLUTION", "EVOLUTIONS", "ABILITIES", "MOVE", "NEW MOVES", "NEW ABILITIES"}: return False
    if u.startswith("LV."): return False
    if re.match(r"^LV\s*\d+$", u): return False
    if re.match(r"^\d+$", u): return False
    if "REGIONAL" in u or "VARIANT" in u: return False
    if not re.search(r"[A-Z]", u): return False
    # Reject evolution-target placeholder cells like "HITMONLEE ➡️" or
    # "⭐ JYNX ➡️" — these appear in evolution columns of other species'
    # blocks and would otherwise create empty species records that
    # overwrite real data via all_species.update(sp).
    if any(arrow in u for arrow in ("➡️", "➡", "➜", "→", "⮕", "->")):
        return False
    return True

DEX_ALIASES = {
    "CLAYDAL": "CLAYDOL",
    "DUDUNSPARS": "DUDUNSPARCE",
    # Battle Bond forms in the Pokédex sheet are listed as "B.B. <Name>"; the
    # species sheets call them "<NAME> (BATTLE BOND)". Map the dex entries so
    # the species canon resolves to the right dex#.
    "BBPLUSLE":   "PLUSLEBATTLEBOND",
    "BBMINUN":    "MINUNBATTLEBOND",
    "BBKECLEON":  "KECLEONBATTLEBOND",
    "BBMAWILE":   "MAWILEBATTLEBOND",
    "BBBLAZIKEN": "BLAZIKENBATTLEBOND",
}

# Pokédex sheet display fixes (typos): canon(dex name) -> proper name.
DEX_NAME_FIXES = {
    "DUDUNSPARS": "Dudunsparce",
}

# Pokémon Showdown's sprite slug doesn't always match our derived slug. Map
# canon(name) -> the literal slug Showdown serves.
SPRITE_SLUG_FIXES = {
    "SANDYSHOCK":  "sandyshocks",
    "SANDYSHOCKS": "sandyshocks",
    "DUDUNSPARS":  "dudunsparce",
    "DUDUNSPARCE": "dudunsparce",
    # Mainline regional forms — the docs list only the regional form so the
    # base slug should serve the regional Showdown sprite.
    "GRIMER":         "grimer-alola",
    "MUK":            "muk-alola",
    "VOLTORB":        "voltorb-hisui",
    "ELECTRODE":      "electrode-hisui",
    "TYPHLOSION":     "typhlosion-hisui",
    # sprite_slug() strips parentheticals (e.g. "(Galar)"), so regional forms
    # whose key includes the region tag need an explicit override here.
    "FARFETCHDGALAR": "farfetchd-galar",
}

# Species that are flagged with ⭐ in the docs but are actually mainline-game
# regional/Hisuian forms, not Etrian Variants. Don't ribbon them as Variant.
NOT_ETRIAN_VARIANT = {
    "GRIMER", "MUK",
    "VOLTORB", "ELECTRODE",
    "TYPHLOSION",
}

# Custom forms (Etrian Variants, Battle Bond, etc.) that should be tagged as variants
# for filtering purposes even if they're not marked with ⭐ in the source sheets.
CUSTOM_VARIANTS = {
    "GOROCHU",  # Custom Etrian Variant form
    "PLUSLEBATTLEBOND", "MINUNBATTLEBOND", "KECLEONBATTLEBOND",
    "MAWILEBATTLEBOND", "BLAZIKENBATTLEBOND",  # Battle Bond custom forms
    "GOLEM",  # Custom Odyssey variant sprite
}

def nidoran_canon(name: str) -> str:
    n = (name or "").upper()
    if "NIDORAN" in n:
        if "(F)" in n or "♀" in n: return "NIDORANF"
        if "(M)" in n or "♂" in n: return "NIDORANM"
    return canon(name)

def item_slug(name: str) -> str:
    s = (name or "").strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = s.lower()
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")


# -----------------------------------------------------------------------------
# Stat sheet parser (now tracks evolution family bands)
# -----------------------------------------------------------------------------

def parse_stat_sheet(ws, sheet_name):
    """Return (species_dict, stats_dict, family_bands) all keyed by canon-name."""
    max_row = ws.max_row or 2000
    raw = list(ws.iter_rows(min_row=1, max_row=max_row, max_col=17, values_only=True))
    rows = [list(r) + [None] * (17 - len(r)) for r in raw]

    species = {}
    # --- pass 1: species data on cols A-B, D-E, G-H ---
    for col_start in (0, 3, 6):
        i = 0
        while i < len(rows):
            name_cell = rows[i][col_start]
            if is_species_header(name_cell):
                display_name = str(name_cell).strip()
                key = nidoran_canon(display_name)
                entry = {
                    "display_name": display_name,
                    "is_variant": "⭐" in display_name,
                    "source_sheet": sheet_name,
                    "_band_row": i,
                    "_band_col": col_start,
                }
                r = i + 1
                while r < len(rows) and r < i + 6:
                    label = rows[r][col_start]
                    value = rows[r][col_start + 1]
                    if isinstance(label, str):
                        lab = label.strip().rstrip(":").upper()
                        if lab == "TYPE":
                            entry["types"] = [t.strip() for t in str(value).split("/")] if value else []
                        elif lab == "ABILITY":
                            entry["abilities"] = [a.strip() for a in str(value).split("/")] if value else []
                        elif lab == "EVOLUTION":
                            v = str(value).strip() if value else ""
                            entry["evolves_at"] = None if v in ("", "/") else v
                        elif lab == "MOVES":
                            break
                    r += 1
                if r < len(rows):
                    mv_here = rows[r][col_start]
                    if isinstance(mv_here, str) and mv_here.strip().upper() == "MOVES":
                        r += 1
                moves = []
                while r < len(rows):
                    lvl = rows[r][col_start]
                    mv  = rows[r][col_start + 1]
                    if lvl is None and mv is None:
                        break
                    if is_species_header(lvl):
                        break
                    if isinstance(lvl, str) and lvl.strip().upper().startswith("LV."):
                        level_str = lvl.strip().replace("LV.", "").replace("LV ", "").strip()
                        try:
                            level = int(level_str)
                        except ValueError:
                            level = level_str
                        moves.append({"level": level, "name": str(mv).strip() if mv else ""})
                    r += 1
                entry["moves"] = moves
                species[key] = entry
                i = r
            else:
                i += 1

    # --- pass 2: stat blocks in col J (index 9) ---
    stats = {}
    i = 0
    while i < len(rows):
        nc = rows[i][9]
        if is_species_header(nc):
            name = str(nc).strip()
            key = nidoran_canon(name)
            odyssey_row = rows[i + 2] if i + 2 < len(rows) else None
            vanilla_row = rows[i + 3] if i + 3 < len(rows) else None

            def read_stats(r):
                if not r: return None
                names_ = ["hp", "atk", "def", "spa", "spd", "spe", "total"]
                vals = r[9:16]
                out = {}
                for n, v in zip(names_, vals):
                    if v is None: continue
                    try: out[n] = int(float(v))
                    except (TypeError, ValueError): pass
                return out or None

            stats[key] = {
                "odyssey": read_stats(odyssey_row),
                "vanilla": read_stats(vanilla_row),
            }
            i += 4
        else:
            i += 1

    # --- pass 3: evolution bands (species on same row in cols A, D, G = same family) ---
    family_bands = []   # each band: { stages: [canon_key,...], branches: [{from_key, to_key, condition}] }
    seen = set()
    for i_band in range(len(rows)):
        a = rows[i_band][0]; d = rows[i_band][3]; g = rows[i_band][6]
        if not is_species_header(a) and not is_species_header(d) and not is_species_header(g):
            continue
        # Avoid reprocessing: require col A species header here (band anchor)
        if not is_species_header(a):
            continue
        if i_band in seen:
            continue
        seen.add(i_band)
        stages = []
        if is_species_header(a):
            stages.append(nidoran_canon(str(a).strip()))
        if is_species_header(d):
            stages.append(nidoran_canon(str(d).strip()))
        if is_species_header(g):
            # Special case: "EVOLUTION" sub-header in D → G is a branch target, not a sequential stage
            # Also handle "ESPEON ➡️" etc.
            g_clean = re.sub(r"[➡➜→⮕]+\s*", "", str(g).strip()).strip()
            if g_clean:
                stages.append(nidoran_canon(g_clean))
        # --- detect branching EVOLUTION panel (under col D or G) ---
        # Pattern: two rows below species row contains "EVOLUTION" label in some col,
        # and subsequent rows list "<condition> | <target-species>" pairs.
        branches = []
        for branch_col in (3, 6):
            r2 = i_band + 1
            if r2 < len(rows) and isinstance(rows[r2][branch_col], str) \
               and rows[r2][branch_col].strip().upper() == "EVOLUTION":
                source_key = stages[0]  # branches always originate from col-A species
                k = r2 + 1
                while k < len(rows) and k < r2 + 10:
                    cond = rows[k][branch_col]
                    tgt  = rows[k][branch_col + 1]
                    if cond is None and tgt is None:
                        break
                    if not isinstance(cond, str) or not isinstance(tgt, str):
                        k += 1; continue
                    tgt_name = re.sub(r"[➡➜→⮕]+\s*", "", tgt.strip()).strip()
                    if tgt_name:
                        branches.append({
                            "from": source_key,
                            "to": nidoran_canon(tgt_name),
                            "to_display": tgt_name,
                            "condition": cond.strip(),
                        })
                    k += 1
        family_bands.append({"stages": stages, "branches": branches, "_row": i_band})

    return species, stats, family_bands

# -----------------------------------------------------------------------------
# Pokédex index
# -----------------------------------------------------------------------------

def load_pokedex_index(wb):
    ws = wb["Pokédex"]
    out = {}
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or len(row) < 3: continue
        _, dex, name = row[0], row[1], row[2]
        if dex and name and isinstance(name, str):
            try:
                dex_str = str(int(dex)).zfill(3) if isinstance(dex, (int, float)) else str(dex).strip().zfill(3)
            except Exception:
                continue
            key = nidoran_canon(name)
            out[key] = {"dex": dex_str, "name": name.strip()}
            raw = canon(name)
            if raw in DEX_ALIASES:
                out[DEX_ALIASES[raw]] = {"dex": dex_str, "name": name.strip()}
    return out

# -----------------------------------------------------------------------------
# Wild encounters (Pokémon sheet)
# -----------------------------------------------------------------------------

HABITAT_SET = {
    "TALL GRASS", "HEADBUTT", "SURF", "FISHING", "OLD ROD", "GOOD ROD", "SUPER ROD",
    "CAVE", "WATER", "ROCK SMASH", "FLOOR", "EVENT", "INTERIOR", "GIFT", "CAVE INTERIOR",
    "DUNGEON", "DEEP WATER", "SHALLOW WATER", "NIGHT", "DAY", "MORNING",
    "LEVITATING", "DESERT", "SAND", "ROCKS", "SNOW", "TRADE",
}

def decode_level(v):
    # Excel auto-formats ranges like "4-2" (levels 2–4) as dates (2022-04-02).
    # Recover by extracting day/month and sorting them into lo-hi order.
    if v is None or v == "": return None
    if isinstance(v, datetime):
        lo, hi = sorted([v.day, v.month])
        return f"{lo}-{hi}" if lo != hi else str(lo)
    if isinstance(v, (int, float)):
        return str(int(v)) if float(v).is_integer() else str(v)
    return str(v).strip()

def parse_wild_encounters(wb):
    ws = wb["Pokémon"]
    max_col = 14
    rows = [list(r) + [None] * (max_col - len(r))
            for r in ws.iter_rows(max_col=max_col, values_only=True)]
    out = {}
    for col in (0, 4, 8, 12):
        location = None
        habitat = None
        in_data = False
        for row in rows:
            a = row[col]
            b = row[col + 1] if col + 1 < max_col else None
            c = row[col + 2] if col + 2 < max_col else None
            if isinstance(a, str) and a.strip().upper() == "POKÉMON" \
                    and isinstance(b, str) and b.strip().upper() == "LEVEL":
                in_data = True
                continue
            if in_data:
                if a is None or a == "":
                    in_data = False
                    continue
                mon = str(a).strip()
                out.setdefault(nidoran_canon(mon), []).append({
                    "display_name": mon,
                    "location": location,
                    "habitat": habitat,
                    "level": decode_level(b),
                    "percent": float(c) if isinstance(c, (int, float)) else (None if c in (None, "") else str(c)),
                })
            else:
                if isinstance(a, str) and a.strip() and a.strip() == a.strip().upper() \
                        and b in (None, "") and c in (None, ""):
                    txt = a.strip()
                    if txt in HABITAT_SET or any(txt.startswith(h) for h in HABITAT_SET):
                        habitat = txt
                    else:
                        location = txt
                        habitat = None
    return out

# -----------------------------------------------------------------------------
# New Moves & Abilities — full extractor with type/category icon decoding
# -----------------------------------------------------------------------------

# md5(image_bytes)[:10] → label, derived once by visual inspection of the
# embedded icons in the New Moves & Abilities sheet. Type/category cells in
# the workbook are images, not text — without this map they read as None.
TYPE_ICON_HASHES = {
    "acedc2bd1d": "Normal",   "a651c0dc46": "Electric", "cc6cb7014f": "Water",
    "5c42af5ab1": "Fire",     "cfe79204eb": "Grass",    "fe7e013ede": "Psychic",
    "7e6817ca40": "Poison",   "f4bd1b822a": "Dark",     "0d75b05da7": "Ground",
    "5db711ef2f": "Fighting", "3c86284e71": "Ice",      "7e1f45411f": "Steel",
    "475797df1a": "Rock",     "411514b790": "Bug",      "1f57507c9f": "Flying",
    "c824276960": "Dragon",   "ccd8a2bd27": "Ghost",    "705cc39322": "Aether",
}
CATEGORY_ICON_HASHES = {
    "67f4636950": "Status",
    "110a105c8e": "Special",
    "db7984cc21": "Physical",
}

# Maps the workbook's section header → (kind, label). The label appears on
# each entry so the UI can distinguish e.g. brand-new moves from reworked
# vanilla ones.
_MOVE_SECTIONS = {
    "NEW MOVES":              ("move",    "new"),
    "AETHER-TYPE MOVES":      ("move",    "aether"),
    "BUFFED/REWORKED MOVES":  ("move",    "reworked"),
    "NEW ABILITIES":          ("ability", "new"),
    "BUFFED ABILITIES":       ("ability", "reworked"),
}

# Field labels we must skip when scanning for entry-name candidates.
_FIELD_LABELS = {"TYPE", "CATEGORY", "POWER", "ACCURACY", "PP", "EFFECT"}


def _build_icon_map(ws):
    """Return {(anchor_row, anchor_col): label} for every type/category icon
    embedded in the worksheet. Coords are 0-indexed (openpyxl image anchors)."""
    import hashlib
    out = {}
    for img in getattr(ws, "_images", []):
        data = img._data() if callable(img._data) else img._data
        h = hashlib.md5(data).hexdigest()[:10]
        label = TYPE_ICON_HASHES.get(h) or CATEGORY_ICON_HASHES.get(h)
        if not label:
            continue
        a = img.anchor._from
        out[(a.row, a.col)] = label
    return out


def _norm_num(v):
    """Normalize a numeric workbook cell. Returns int, str (e.g. '???'), or None."""
    if v is None or v == "" or v == "-":
        return None
    if isinstance(v, (int, float)):
        return int(v) if float(v).is_integer() else v
    return str(v).strip()


def parse_moves_and_abilities(xlsx_path):
    """Parse the 'New Moves & Abilities' sheet. Returns (moves, abilities)
    dicts keyed by the entry's site slug.

    Five sections in the sheet are recognized:
      - NEW MOVES, AETHER-TYPE MOVES, BUFFED/REWORKED MOVES → moves
      - NEW ABILITIES, BUFFED ABILITIES                     → abilities

    Move blocks are 8 rows tall (name, type, category, power, accuracy, pp,
    'Effect' label, effect text) with up to 6 entries per band at a 3-col
    stride. Ability blocks are 2 rows (name + effect text)."""
    # Must NOT be read_only — image objects aren't exposed in read-only mode.
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb["New Moves & Abilities"]
        icon_map = _build_icon_map(ws)
        rows = [list(r) + [None] * (20 - len(r))
                for r in ws.iter_rows(max_col=20, values_only=True)]

        moves, abilities = {}, {}
        section_kind = None
        section_label = None
        name_cols = (1, 4, 7, 10, 13, 16)

        for i, row in enumerate(rows):
            first = row[0]
            if isinstance(first, str):
                t = first.strip().upper()
                if t in _MOVE_SECTIONS:
                    section_kind, section_label = _MOVE_SECTIONS[t]
                    continue
            if section_kind is None:
                continue

            # Cells that look like ALL-CAPS entry names (filters out the
            # field-label rows and the lower-case effect-text rows).
            candidates = [(c, row[c]) for c in name_cols
                          if isinstance(row[c], str) and row[c].strip()
                          and row[c].strip() == row[c].strip().upper()
                          and row[c].strip() not in _FIELD_LABELS]
            if not candidates:
                continue

            if section_kind == "ability":
                # 2-row block: name + effect text directly below.
                if i + 1 >= len(rows):
                    continue
                for col, name in candidates:
                    nm = name.strip()
                    slug = slugify(nm)
                    eff = rows[i + 1][col]
                    abilities[slug] = {
                        "name": nm,
                        "slug": slug,
                        "effect": str(eff).strip() if eff else None,
                        "kind": section_label,         # 'new' | 'reworked'
                        "is_custom": True,
                    }
            else:  # 'move' — 8-row block, confirmed by next row's TYPE label
                if i + 1 >= len(rows):
                    continue
                confirmed = [(c, n) for c, n in candidates
                             if isinstance(rows[i + 1][c], str)
                             and rows[i + 1][c].strip().upper() == "TYPE"]
                for col, name in confirmed:
                    nm = name.strip()
                    slug = slugify(nm)
                    type_lbl = icon_map.get((i + 1, col + 1))
                    cat_lbl  = icon_map.get((i + 2, col + 1))
                    eff = rows[i + 7][col] if i + 7 < len(rows) else None
                    moves[slug] = {
                        "name": nm,
                        "slug": slug,
                        "type": type_lbl,
                        "category": cat_lbl,
                        "power":    _norm_num(rows[i + 3][col + 1]) if i + 3 < len(rows) else None,
                        "accuracy": _norm_num(rows[i + 4][col + 1]) if i + 4 < len(rows) else None,
                        "pp":       _norm_num(rows[i + 5][col + 1]) if i + 5 < len(rows) else None,
                        "effect":   str(eff).strip() if eff else None,
                        "kind": section_label,         # 'new' | 'aether' | 'reworked'
                        "is_custom": True,
                    }
        return moves, abilities
    finally:
        wb.close()


# -----------------------------------------------------------------------------
# Type chart — extracted from cell fill colors on the Type Chart sheet.
# -----------------------------------------------------------------------------

# Legend colors verified against the workbook's own legend rows (R22/R24/R26).
_TYPE_FILL_TO_MULT = {
    "FFE06666": 0,    # red    — no effect
    "FFFFE599": 0.5,  # yellow — not very effective
    "FF93C47D": 2,    # green  — super effective
}


def parse_type_chart(xlsx_path):
    """Return {attacker: {defender: multiplier}} for every non-1x interaction.
    Reads the Type Chart sheet's cell fills since the multipliers themselves
    aren't stored as text. Aether is the 18th type and only differs from the
    standard chart in those rows/cols."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb["Type Chart"]
        defenders = [ws.cell(2, c).value for c in range(3, 21)]
        attackers = [ws.cell(r, 2).value for r in range(3, 21)]
        chart = {}
        for ri, atk in enumerate(attackers, start=3):
            chart[atk] = {}
            for ci, deff in enumerate(defenders, start=3):
                f = ws.cell(ri, ci).fill
                if not (f and f.patternType == "solid" and f.fgColor):
                    continue
                rgb = (f.fgColor.rgb or "").upper()
                if rgb in _TYPE_FILL_TO_MULT:
                    chart[atk][deff] = _TYPE_FILL_TO_MULT[rgb]
        return chart
    finally:
        wb.close()


# -----------------------------------------------------------------------------
# PokeAPI baseline fetcher (with on-disk cache)
# -----------------------------------------------------------------------------

POKEAPI_BASE  = "https://pokeapi.co/api/v2"
POKEAPI_CACHE = ROOT / "cache" / "pokeapi_cache"


def _pokeapi_slug(name):
    """Convert a workbook entry name to a PokeAPI URL slug."""
    s = (name or "").strip().lower()
    s = s.replace("\u2019", "").replace("'", "")     # apostrophes
    s = s.replace("\u2014", "-").replace("\u2013", "-")
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s


# Manual one-off fixes for docs name typos / smashed-together names that
# refer to known mainline moves. These map (canon-of-docs-name) -> the
# real PokeAPI slug.
POKEAPI_NAME_FIXES = {
    # Typos / mis-spellings in the source spreadsheet
    "EARTQUAKE":      "earthquake",
    "DEFENCECURL":    "defense-curl",
    "ATONISH":        "astonish",
    "KEENEYES":       "keen-eye",
    # Hi vs High
    "HIHORSEPOWER":   "high-horsepower",
    "HIJUMPKICK":     "high-jump-kick",
    # "First Press" -> First Impression in mainline
    "FIRSTPRESS":     "first-impression",
}


# Cached canonical-form -> real PokeAPI slug index for moves and abilities.
# Built lazily on first miss so we only do the network round-trip when we
# actually need to recover from a smashed-together docs name.
_POKEAPI_INDEX_CACHE = {}  # kind -> {CANON: real-slug}


def _pokeapi_canon_index(kind):
    """Return {canon(real_slug): real_slug} for every entry of `kind`.
    `kind` is "move" or "ability". Cached on disk and in-memory."""
    if kind in _POKEAPI_INDEX_CACHE:
        return _POKEAPI_INDEX_CACHE[kind]
    POKEAPI_CACHE.mkdir(parents=True, exist_ok=True)
    cache_file = POKEAPI_CACHE / f"_index_{kind}.json"
    if cache_file.exists():
        names = json.loads(cache_file.read_text())
    else:
        import urllib.request
        url = f"{POKEAPI_BASE}/{kind}?limit=10000"
        req = urllib.request.Request(url, headers={
            "User-Agent": "PokemonOdysseyDocs/1.0 (+github.com/nishxpatel/Pokemon-Odyssey-Docs-App)",
            "Accept": "application/json",
        })
        with urllib.request.urlopen(req, timeout=30) as r:
            data = json.loads(r.read().decode())
        names = [e["name"] for e in (data.get("results") or [])]
        cache_file.write_text(json.dumps(names))
    idx = {canon(n.replace("-", "")): n for n in names}
    _POKEAPI_INDEX_CACHE[kind] = idx
    return idx


def _resolve_pokeapi(kind, name):
    """Try to fetch a PokeAPI entry of `kind` for `name`, with multiple
    fallback strategies for docs names that smash words together
    (Thundershock, Lightningrod, Compoundeyes...). Returns
    (real_slug, data) — the canonical PokeAPI slug paired with the raw
    response dict — or (None, None)."""
    # 1) Manual fix table (typos, hi/high, etc.)
    fix = POKEAPI_NAME_FIXES.get(canon(name))
    if fix:
        d = _fetch_pokeapi(kind, fix)
        if d: return fix, d
    # 2) Direct slug
    slug = _pokeapi_slug(name)
    if slug:
        d = _fetch_pokeapi(kind, slug)
        if d: return slug, d
    # 3) Canon-form lookup against the full PokeAPI listing — handles
    # smashed-together names by collapsing both sides to letters+digits.
    try:
        idx = _pokeapi_canon_index(kind)
    except Exception:
        idx = {}
    real = idx.get(canon(name))
    if real and real != slug:
        d = _fetch_pokeapi(kind, real)
        if d: return real, d
    return None, None


def _fetch_pokeapi(kind, slug):
    """kind: 'move' or 'ability'. Returns the raw dict, or None on 404.
    Caches every response (including 404s as 'null') under pokeapi_cache/."""
    import time
    import urllib.request, urllib.error
    cache_dir = POKEAPI_CACHE / f"{kind}s"
    cache_dir.mkdir(parents=True, exist_ok=True)
    cache_file = cache_dir / f"{slug}.json"
    if cache_file.exists():
        txt = cache_file.read_text()
        return json.loads(txt) if txt.strip() != "null" else None
    url = f"{POKEAPI_BASE}/{kind}/{slug}"
    req = urllib.request.Request(url, headers={
        "User-Agent": "PokemonOdysseyDocs/1.0 (+github.com/nishxpatel/Pokemon-Odyssey-Docs-App)",
        "Accept": "application/json",
    })
    try:
        with urllib.request.urlopen(req, timeout=20) as r:
            data = json.loads(r.read().decode())
    except urllib.error.HTTPError as e:
        if e.code == 404:
            cache_file.write_text("null")
            return None
        raise
    cache_file.write_text(json.dumps(data))
    time.sleep(0.05)   # be polite
    return data


def _english_short_effect(data):
    """Best-effort English effect text from a PokeAPI move/ability response."""
    for e in (data.get("effect_entries") or []):
        if (e.get("language") or {}).get("name") == "en":
            return (e.get("short_effect") or e.get("effect") or "").strip() or None
    # Some moves only have flavor_text_entries
    for e in (data.get("flavor_text_entries") or []):
        if (e.get("language") or {}).get("name") == "en":
            return (e.get("flavor_text") or "").replace("\n", " ").replace("\f", " ").strip() or None
    return None


def baseline_move(name):
    """Fetch a baseline (non-custom) move from PokeAPI. Returns a normalized
    entry matching parse_moves_and_abilities output, or None if unknown.
    Stores the entry under the canonical PokeAPI slug so docs name variants
    (e.g. 'Bubblebeam' and 'Bubble Beam') collapse into one entry."""
    real_slug, data = _resolve_pokeapi("move", name)
    if not data:
        return None
    # Pretty display name comes from the PokeAPI slug, not the docs spelling,
    # so the move list shows 'Bubble Beam' not 'BUBBLEBEAM'.
    display = real_slug.replace("-", " ").upper() if real_slug else name.strip().upper()
    cat = (data.get("damage_class") or {}).get("name")
    return {
        "name": display,
        "slug": real_slug,
        "type": ((data.get("type") or {}).get("name") or "").title() or None,
        "category": cat.title() if cat else None,
        "power": data.get("power"),
        "accuracy": data.get("accuracy"),
        "pp": data.get("pp"),
        "effect": _english_short_effect(data),
        "kind": "baseline",
        "is_custom": False,
    }


def baseline_ability(name):
    """Fetch a baseline ability from PokeAPI. Returns normalized entry or None.
    Stored under the canonical PokeAPI slug to dedupe docs name variants."""
    real_slug, data = _resolve_pokeapi("ability", name)
    if not data:
        return None
    display = real_slug.replace("-", " ").upper() if real_slug else name.strip().upper()
    return {
        "name": display,
        "slug": real_slug,
        "effect": _english_short_effect(data),
        "kind": "baseline",
        "is_custom": False,
    }

# -----------------------------------------------------------------------------
# Items: shops, pickup, gathering/mining, TM, tutors
# -----------------------------------------------------------------------------

def _is_all_caps_header(v):
    return isinstance(v, str) and v.strip() and v.strip() == v.strip().upper() and re.search(r"[A-Z]", v)

def add_item_source(items, name, source):
    if not name or not isinstance(name, str): return
    n = name.strip()
    if not n: return
    key = item_slug(n)
    if not key: return
    entry = items.setdefault(key, {"name": n, "slug": key, "sources": []})
    # Keep the prettiest form of the name
    if len(n) < len(entry["name"]) and n.istitle():
        entry["name"] = n
    entry["sources"].append(source)

def parse_items_sheet(wb, items):
    """`Items` sheet: locations with item lists. Layout is varied; we scan for
    section headers in col A (location), keep state, and grab paired (name, note)
    cells across the row."""
    ws = wb["Items"]
    rows = [list(r) + [None] * (10 - len(r))
            for r in ws.iter_rows(max_col=10, values_only=True)]
    location = None
    habitat = None
    for row in rows:
        a = row[0]
        # header: all caps, nothing else on the row
        if _is_all_caps_header(a) and all(x in (None, "") for x in row[1:]):
            location = a.strip()
            habitat = None
            continue
        # sub-header like "GATHERING", "MINING", "ITEMS"
        if _is_all_caps_header(a) and row[1] in (None, ""):
            habitat = a.strip()
            continue
        # data row: item in col B, type/note in col C
        name = row[1]
        note = row[2]
        if isinstance(name, str) and name.strip():
            add_item_source(items, name, {
                "kind": "location",
                "location": location,
                "habitat": habitat,
                "note": note if isinstance(note, str) else None,
            })
        # also parallel block in cols F/G/H
        name2 = row[5]; note2 = row[6]
        if isinstance(name2, str) and name2.strip():
            add_item_source(items, name2, {
                "kind": "location",
                "location": location,
                "habitat": habitat,
                "note": note2 if isinstance(note2, str) else None,
            })

def parse_items_shop_sheet(wb, items):
    ws = wb["Items (Shop)"]
    rows = [list(r) + [None] * (17 - len(r))
            for r in ws.iter_rows(max_col=17, values_only=True)]
    shop_name = None
    # LEVEL headers live on a single row; track which col ranges map to which level
    levels_by_col = {}
    for row in rows:
        # Shop-name row: one caps string at col 0, rest blank
        if _is_all_caps_header(row[0]) and all(x in (None, "") for x in row[1:]):
            shop_name = row[0].strip()
            levels_by_col = {}
            continue
        # LEVEL headers row
        level_header = False
        for idx in range(17):
            v = row[idx]
            if isinstance(v, str) and re.match(r"^\s*LEVEL\s*\d+\s*$", v.strip(), re.I):
                level_header = True
                levels_by_col[idx] = v.strip().upper()
        if level_header:
            continue
        # "TREASURES OBTAINED: N" variants
        for idx in range(17):
            v = row[idx]
            if isinstance(v, str) and v.strip().upper().startswith("TREASURES OBTAINED"):
                levels_by_col[idx] = v.strip()
        # data row: item cells near each LEVEL col
        for col, lvl in levels_by_col.items():
            v = row[col]
            if isinstance(v, str) and v.strip() and v.strip().upper() != lvl.upper():
                add_item_source(items, v, {
                    "kind": "shop",
                    "shop": shop_name,
                    "level": lvl,
                })

def parse_items_pickup_sheet(wb, items):
    ws = wb["Items (Pickup)"]
    rows = list(ws.iter_rows(max_col=5, values_only=True))
    for row in rows:
        pct = row[1] if len(row) > 1 else None
        name = row[2] if len(row) > 2 else None
        if isinstance(name, str) and name.strip() and not name.strip().upper() in {"ITEM", "PERCENTAGE"}:
            add_item_source(items, name, {
                "kind": "pickup",
                "percent": pct if isinstance(pct, (int, float)) else None,
            })

def parse_gathering_mining_sheet(wb, items):
    ws = wb["GatheringMining"]
    rows = [list(r) + [None] * (10 - len(r))
            for r in ws.iter_rows(max_col=10, values_only=True)]
    stratum = None
    sub = {0: None, 3: None, 6: None}
    for row in rows:
        a = row[0]
        # Stratum header: single caps value at col 0
        if _is_all_caps_header(a) and all(x in (None, "") for x in row[1:]):
            stratum = a.strip()
            sub = {0: None, 3: None, 6: None}
            continue
        # Sub-header row: caps at cols 0/3/6 like "GATHERING", "MINING"
        maybe = False
        for c in (0, 3, 6):
            v = row[c]
            if isinstance(v, str) and v.strip() and v.strip().upper() == v.strip():
                if any(k in v.upper() for k in ("GATHERING", "MINING", "BERRY", "BERRIES", "FLOWER")):
                    sub[c] = v.strip()
                    maybe = True
        if maybe:
            continue
        # data rows: item strings at cols 0/3/6 (name only, no note)
        for c in (0, 3, 6):
            v = row[c]
            if isinstance(v, str) and v.strip():
                s = v.strip()
                if s.upper() == s and re.search(r"[A-Z]{4,}", s):
                    # uppercase likely header we missed; skip
                    continue
                add_item_source(items, s, {
                    "kind": "gather",
                    "stratum": stratum,
                    "method": sub[c],
                })

def parse_tm_sheet(wb, items):
    ws = wb["TM Location"]
    for row in ws.iter_rows(min_row=4, max_col=5, values_only=True):
        num, move, loc = (row + (None,)*5)[:3]
        if move and isinstance(move, str):
            tm_name = f"TM{int(num):02d} {move.strip()}" if isinstance(num, (int, float)) else f"TM {move.strip()}"
            add_item_source(items, tm_name, {
                "kind": "tm",
                "move": move.strip(),
                "location": loc.strip() if isinstance(loc, str) else None,
            })

def parse_move_tutors_sheet(wb):
    ws = wb["Move Tutors"]
    tutors = []
    for row in ws.iter_rows(min_row=4, max_col=5, values_only=True):
        move, loc = (row + (None,)*5)[1], (row + (None,)*5)[2]
        if isinstance(move, str) and move.strip():
            tutors.append({
                "move": move.strip(),
                "location": loc.strip() if isinstance(loc, str) else None,
            })
    return tutors

# -----------------------------------------------------------------------------
# Variant sprite extraction
# -----------------------------------------------------------------------------

def _clean_originals_to_variants(originals_dir: Path, cleaned_dir: Path) -> None:
    """Strip the solid bg from every PNG in originals_dir into cleaned_dir.

    Strict equality removal — bg is whatever color the corner of each image
    actually is. Pixel-art assets have no anti-aliasing so this will not eat
    sprite pixels even when the body color is visually similar to the bg.
    """
    try:
        from clean_variant_backgrounds import clean_dir
    except ImportError:
        return
    clean_dir(originals_dir, cleaned_dir)


def extract_variant_sprites(xlsx_path: Path, out_dir: Path):
    """Pull embedded PNGs from the 'Etrian Variants' sheet and save under out_dir.

    Returns a mapping: canon(species_or_variant_name) -> {normal: path, shiny: path, variant_name: str}.
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    NS = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a':   'http://schemas.openxmlformats.org/drawingml/2006/main',
        'r':   'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
    }

    # Read sheet data for name rows (we need base-species → row map)
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Etrian Variants"]
    ws.calculate_dimension(force=True)
    sheet_rows = list(ws.iter_rows(max_col=15, values_only=True))
    wb.close()

    # Build block list. A block is 7 rows tall starting at the name row.
    # Block-start detection: row N has all-caps species names in cols 1+, and
    # row N+1 col 0 == "NORMAL".
    blocks = []
    for i in range(len(sheet_rows) - 5):
        r = sheet_rows[i]
        rn = sheet_rows[i + 1]
        if not isinstance(rn[0], str) or rn[0].strip().upper() != "NORMAL":
            continue
        base_names = {}   # col -> base species
        for c in range(1, 15):
            v = r[c]
            if isinstance(v, str) and v.strip() and v.strip() == v.strip().upper():
                base_names[c] = v.strip()
        # Battle Bond blocks have no Odyssey rename (they're already named
        # "<MON> (BATTLE BOND)"), and the row that would normally hold variant
        # names actually contains stray entries from a different concept. Skip
        # variant_names for them so sprites keep the BATTLE BOND-derived slug.
        is_battle_bond_block = any("BATTLE BOND" in n for n in base_names.values())
        # Variant name row at i+4
        variant_names = {}
        if i + 4 < len(sheet_rows) and not is_battle_bond_block:
            rv = sheet_rows[i + 4]
            for c in range(1, 15):
                v = rv[c]
                if isinstance(v, str) and v.strip() and v.strip() == v.strip().upper():
                    variant_names[c] = v.strip()
        blocks.append({
            "name_row": i,         # 0-indexed
            "normal_row": i + 1,
            "shiny_row":  i + 2,
            "variant_row": i + 4,
            "base_names": base_names,
            "variant_names": variant_names,
        })

    # Open xlsx as zip to grab drawings + images
    with zipfile.ZipFile(xlsx_path) as z:
        # Find the right drawing file for "Etrian Variants". We discover by
        # scanning xl/worksheets/_rels and xl/workbook.xml to map sheet → drawing.
        # Simple approach: iterate all drawings and match against images whose
        # anchors land on our block rows.
        drawing_names = [n for n in z.namelist() if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
        # Find the drawing whose anchors fall on our rows; it's the one for this sheet.
        target_drawing = None
        for dname in drawing_names:
            x = ET.fromstring(z.read(dname))
            sample_rows = []
            for anc in x.findall('xdr:oneCellAnchor', NS) + x.findall('xdr:twoCellAnchor', NS):
                frm = anc.find('xdr:from', NS)
                if frm is not None:
                    row_el = frm.find('xdr:row', NS)
                    if row_el is not None:
                        sample_rows.append(int(row_el.text))
                if len(sample_rows) > 5: break
            block_rows_set = {b["normal_row"] for b in blocks} | {b["shiny_row"] for b in blocks}
            if any(r in block_rows_set for r in sample_rows):
                target_drawing = dname
                break
        if target_drawing is None:
            print("  warn: no drawing file matched Etrian Variants sheet", file=sys.stderr)
            return {}

        rels_name = target_drawing.replace("drawings/", "drawings/_rels/") + ".rels"
        rels_root = ET.fromstring(z.read(rels_name))
        rid_to_img = {}
        for rel in rels_root:
            rid = rel.attrib.get("Id")
            tgt = rel.attrib.get("Target")
            if tgt and "media" in tgt:
                rid_to_img[rid] = tgt.replace("../", "xl/")

        d_root = ET.fromstring(z.read(target_drawing))
        # (row, col) → image path
        anchor_to_img = {}
        for anc in d_root.findall('xdr:oneCellAnchor', NS) + d_root.findall('xdr:twoCellAnchor', NS):
            frm = anc.find('xdr:from', NS)
            if frm is None: continue
            col = int(frm.find('xdr:col', NS).text)
            row = int(frm.find('xdr:row', NS).text)
            blip = anc.find('.//a:blip', NS)
            if blip is None: continue
            rid = blip.attrib.get('{%s}embed' % NS['r'])
            img_path = rid_to_img.get(rid)
            if img_path:
                # Don't overwrite if multiple anchors share a cell (shouldn't)
                anchor_to_img.setdefault((row, col), img_path)

        # Now extract per block
        sprite_map = {}   # canon-name → {normal, shiny, variant_display, base_display}
        for blk in blocks:
            for c, base in blk["base_names"].items():
                variant = blk["variant_names"].get(c, base)
                img_n = anchor_to_img.get((blk["normal_row"], c))
                img_s = anchor_to_img.get((blk["shiny_row"], c))
                # Output file name: prefer variant name slug so UI can key by
                # either the base species (⭐ RATTATA) or Odyssey name (TREERAT).
                vslug = slugify(variant)
                normal_path = shiny_path = None
                if img_n:
                    data = z.read(img_n)
                    p = out_dir / f"{vslug}.png"
                    p.write_bytes(data)
                    normal_path = f"assets/variants/{vslug}.png"
                if img_s:
                    data = z.read(img_s)
                    p = out_dir / f"{vslug}-shiny.png"
                    p.write_bytes(data)
                    shiny_path = f"assets/variants/{vslug}-shiny.png"
                rec = {
                    "base_display": base,
                    "variant_display": variant,
                    "normal": normal_path,
                    "shiny": shiny_path,
                }
                # Key by both the base species canon and the variant name canon,
                # so ⭐ RATTATA and TREERAT both resolve.
                sprite_map[canon(base)] = rec
                sprite_map[canon(variant)] = rec
        return sprite_map

# -----------------------------------------------------------------------------
# Evolution chain construction
# -----------------------------------------------------------------------------

def build_evolution_graph(all_species, all_bands):
    """Walk family bands across every stat sheet; build a directed graph
    (from_key -> list of {to_key, condition, kind}).
    Also constructs each species' full "family" (the rooted set of connected stages)."""
    graph = {}   # key -> list of edges
    reverse = {} # key -> list of predecessors (for family lookup)
    edge_keys = {}  # (from, to) -> edge dict (so we can dedupe / upgrade kind)
    def add_edge(a, b, cond, kind):
        if a == b: return
        existing = edge_keys.get((a, b))
        if existing is not None:
            # Branch edges carry more specific conditions than the inferred
            # sequential "stage" edge — prefer them when both exist for the
            # same (from, to) pair (fixes Tyrogue → two Hitmontops dupe).
            if kind == "branch" and existing["kind"] != "branch":
                existing["condition"] = cond
                existing["kind"] = kind
            return
        edge = {"to": b, "condition": cond, "kind": kind}
        edge_keys[(a, b)] = edge
        graph.setdefault(a, []).append(edge)
        reverse.setdefault(b, []).append(a)

    for band in all_bands:
        stages = band["stages"]
        # chain conditions from species entries; only add an edge if source has
        # a non-empty evolves_at — otherwise the band is just a layout grouping
        # of unrelated species (common on 4th Gen / Paradox sheets).
        for i in range(len(stages) - 1):
            src = stages[i]
            dst = stages[i + 1]
            src_sp = all_species.get(src)
            if not src_sp:
                continue
            cond = src_sp.get("evolves_at")
            if not cond:
                continue
            add_edge(src, dst, cond, "stage")
        for b in band["branches"]:
            add_edge(b["from"], b["to"], b["condition"], "branch")

    # Apply explicit branch overrides (EXPLICIT_BRANCHES) and cross-sheet stage
    # evolutions (CROSS_SHEET_EVOS).  Both use the same edge-insertion path;
    # the only difference is the kind tag ("branch" vs "stage"), which the UI
    # uses for display but not for family grouping.
    for from_key, branches in EXPLICIT_BRANCHES.items():
        for to_key, cond in branches:
            add_edge(from_key, to_key, cond, "branch")
    for from_key, evos in CROSS_SHEET_EVOS.items():
        for to_key, cond in evos:
            add_edge(from_key, to_key, cond, "stage")

    return graph, reverse

def family_for(root_key, graph, reverse):
    """Return BFS set of all keys in the evolution family (backwards + forwards)."""
    seen = set([root_key])
    stack = [root_key]
    while stack:
        k = stack.pop()
        for nxt in graph.get(k, []):
            if nxt["to"] not in seen:
                seen.add(nxt["to"]); stack.append(nxt["to"])
        for prev in reverse.get(k, []):
            if prev not in seen:
                seen.add(prev); stack.append(prev)
    return seen

def detect_evolution_items(condition):
    """Return list of item names referenced by an evolution condition string."""
    if not condition: return []
    found = []
    for item in sorted(EVOLUTION_ITEMS, key=len, reverse=True):
        if re.search(rf"\b{re.escape(item)}\b", condition, re.I):
            found.append(item)
    return list(dict.fromkeys(found))  # dedupe, preserve order

# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------

def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Extracting variant sprites to originals dir...", flush=True)
    sprite_map = extract_variant_sprites(STATS_XLSX, ORIGINALS_DIR)
    print(f"  sprites: {sum(1 for v in sprite_map.values() if v.get('normal'))} normal, "
          f"{sum(1 for v in sprite_map.values() if v.get('shiny'))} shiny", flush=True)

    print("Cleaning sprite backgrounds (originals -> variants)...", flush=True)
    _clean_originals_to_variants(ORIGINALS_DIR, VARIANTS_DIR)

    print("Loading Pokémon Stats workbook...", flush=True)
    wb_stats = openpyxl.load_workbook(STATS_XLSX, read_only=True, data_only=True)

    all_species = {}
    all_stats = {}
    all_bands = []
    for name in STAT_SHEETS:
        print(f"  parsing sheet {name!r}", flush=True)
        sp, st, bands = parse_stat_sheet(wb_stats[name], name)
        all_species.update(sp)
        all_stats.update(st)
        all_bands.extend(bands)
    dex_index = load_pokedex_index(wb_stats)
    wb_stats.close()

    for key, val in EVOLVES_AT_OVERRIDES.items():
        if key in all_species:
            all_species[key]["evolves_at"] = val

    # Reopen non-readonly to access embedded type/category icons.
    custom_moves, custom_abilities = parse_moves_and_abilities(STATS_XLSX)
    type_chart = parse_type_chart(STATS_XLSX)

    print("Loading Wild encounters workbook...", flush=True)
    wb_wild = openpyxl.load_workbook(WILD_XLSX, read_only=True, data_only=True)
    locations = parse_wild_encounters(wb_wild)
    items = {}
    parse_items_sheet(wb_wild, items)
    parse_items_shop_sheet(wb_wild, items)
    parse_items_pickup_sheet(wb_wild, items)
    parse_gathering_mining_sheet(wb_wild, items)
    parse_tm_sheet(wb_wild, items)
    tutors = parse_move_tutors_sheet(wb_wild)
    wb_wild.close()

    # Build evolution graph
    graph, reverse = build_evolution_graph(all_species, all_bands)

    print(f"Species parsed: {len(all_species)}", flush=True)
    print(f"Stat blocks parsed: {len(all_stats)}", flush=True)
    print(f"Dex entries: {len(dex_index)}", flush=True)
    print(f"Custom moves: {len(custom_moves)} | Custom abilities: {len(custom_abilities)}", flush=True)
    print(f"Pokémon with wild locations: {len(locations)}", flush=True)
    print(f"Evolution edges: {sum(len(v) for v in graph.values())}", flush=True)
    print(f"Items: {len(items)}", flush=True)

    # -- assemble move/ability indexes (custom + PokeAPI baseline) --
    # Collect every move/ability name referenced across all species.
    referenced_moves = set()
    referenced_abilities = set()
    for sp in all_species.values():
        for m in sp.get("moves", []):
            nm = (m.get("name") or "").strip()
            if nm:
                referenced_moves.add(nm)
        for a in sp.get("abilities", []):
            if a and a.strip():
                referenced_abilities.add(a.strip())

    move_index = {}     # slug -> entry (custom or baseline) with used_by[]
    ability_index = {}
    for m in custom_moves.values():
        move_index[m["slug"]] = dict(m, used_by=[])
    for a in custom_abilities.values():
        ability_index[a["slug"]] = dict(a, used_by=[])

    print("Backfilling baseline data from PokeAPI (cached)...", flush=True)
    # Aliases map a docs-derived slug (slugify(name)) to the canonical
    # PokeAPI slug used as the index key. This is how two docs name
    # variants (e.g. 'Mega Horn' / 'Megahorn' / 'mega-horn' / 'megahorn')
    # collapse to a single move entry while species-level links still
    # resolve from either spelling.
    move_alias = {}
    ability_alias = {}
    fetched_m = fetched_a = missing_m = missing_a = 0
    for nm in sorted(referenced_moves):
        docs_slug = slugify(nm)
        if docs_slug in move_index:
            continue
        # If this docs spelling already aliases to a stored entry, skip.
        if move_alias.get(docs_slug) in move_index:
            continue
        base = baseline_move(nm)
        if not base:
            missing_m += 1
            continue
        real_slug = base["slug"]
        if real_slug not in move_index:
            move_index[real_slug] = dict(base, used_by=[])
            fetched_m += 1
        if real_slug != docs_slug:
            move_alias[docs_slug] = real_slug
    for nm in sorted(referenced_abilities):
        docs_slug = slugify(nm)
        if docs_slug in ability_index:
            continue
        if ability_alias.get(docs_slug) in ability_index:
            continue
        base = baseline_ability(nm)
        if not base:
            missing_a += 1
            continue
        real_slug = base["slug"]
        if real_slug not in ability_index:
            ability_index[real_slug] = dict(base, used_by=[])
            fetched_a += 1
        if real_slug != docs_slug:
            ability_alias[docs_slug] = real_slug
    print(f"  fetched {fetched_m} moves + {fetched_a} abilities from PokeAPI; "
          f"{missing_m} moves + {missing_a} abilities unresolved", flush=True)
    print(f"  aliased {len(move_alias)} move name variants, "
          f"{len(ability_alias)} ability name variants", flush=True)

    def resolve_move_slug(s):
        if s in move_index: return s
        if s in move_alias: return move_alias[s]
        return None

    def resolve_ability_slug(s):
        if s in ability_index: return s
        if s in ability_alias: return ability_alias[s]
        return None

    # -- merge into final Pokédex array --
    merged = []
    unmatched_species = []

    def strip_form(k):
        for suffix in ("GALAR", "ALOLA", "HISUI", "PALDEA", "BATTLEBOND"):
            if k.endswith(suffix) and len(k) > len(suffix):
                return k[:-len(suffix)]
        return k

    for key, sp in all_species.items():
        stats = all_stats.get(key)
        dex = dex_index.get(key) or dex_index.get(strip_form(key))
        if not dex:
            unmatched_species.append(sp["display_name"])
        ev = sp.get("evolves_at")
        ev_level = None
        ev_note = None
        if ev:
            m = re.match(r"LV\.?\s*(\d+)", ev.upper())
            if m:
                ev_level = int(m.group(1))
                ev_note = ev if "/" in ev else None   # hybrid conditions still useful
            else:
                ev_note = ev

        # Evolution targets (from graph, direct successors)
        targets = []
        for edge in graph.get(key, []):
            tgt_sp = all_species.get(edge["to"])
            tgt_dex = dex_index.get(edge["to"]) or dex_index.get(strip_form(edge["to"]))
            tgt_display = tgt_sp["display_name"] if tgt_sp else (tgt_dex["name"] if tgt_dex else edge["to"].title())
            targets.append({
                "to_key": edge["to"],
                "to_name": tgt_display,
                "to_slug": slugify(tgt_display),
                "condition": edge["condition"],
                "kind": edge["kind"],
                "items": detect_evolution_items(edge["condition"]),
            })

        # Full family (all connected species)
        fam_keys = family_for(key, graph, reverse)
        family = []
        for k in fam_keys:
            sp2 = all_species.get(k)
            dex2 = dex_index.get(k) or dex_index.get(strip_form(k))
            disp = sp2["display_name"] if sp2 else (dex2["name"] if dex2 else k.title())
            dex2_canon = canon(dex2["name"]) if dex2 else None
            fam_display = DEX_NAME_FIXES.get(dex2_canon, dex2["name"] if dex2 else disp)
            fam_sprite = SPRITE_SLUG_FIXES.get(k) or SPRITE_SLUG_FIXES.get(dex2_canon) or sprite_slug(fam_display)
            base_only2 = disp.replace("⭐", "").strip()
            family.append({
                "key": k,
                "name": disp,
                "slug": slugify(disp),
                "dex": dex2["dex"] if dex2 else None,
                "sprite_slug": fam_sprite,
                "is_variant": ("⭐" in disp) and canon(base_only2) not in NOT_ETRIAN_VARIANT,
            })
        family.sort(key=lambda x: (0, int(x["dex"])) if x["dex"] else (1, 9999))

        # Event / gift flag: all non-habitat wild entries (tall grass etc.) absent; only EVENT/GIFT
        locs = locations.get(key, [])
        is_event = bool(locs) and all((l["habitat"] or "").upper() in ("EVENT", "GIFT", "TRADE") for l in locs)
        has_wild = bool(locs) and not is_event
        # Identify any items referenced by this species' own evolution condition
        evolution_items = detect_evolution_items(ev)

        # Reclassify species that the docs flag with ⭐ but are actually mainline
        # regional forms (Alolan Grimer/Muk, Hisuian Voltorb line, Hisuian
        # Typhlosion). They keep the ⭐ in display_name but lose the variant flag
        # so the UI ribbon doesn't claim they're Etrian Variants.
        base_only = sp["display_name"].replace("⭐", "").strip()
        is_variant = (sp["is_variant"] and canon(base_only) not in NOT_ETRIAN_VARIANT) or canon(key) in CUSTOM_VARIANTS
        # Battle Bond is a separate custom-form concept from Etrian Variants —
        # tag it independently so the UI can show its own badge and use the
        # custom sprite the workbook provides.
        is_battle_bond = "(BATTLE BOND)" in sp["display_name"].upper()

        # Variant sprite path. Comes from the Etrian Variants sheet for true
        # Etrian Variants and Battle Bond forms, or from a manually-dropped
        # file in site/assets/variants/<slug>.png for custom-art species like
        # Gorochu.
        variant_sprite = None
        if is_variant or is_battle_bond:
            rec = sprite_map.get(canon(base_only)) or sprite_map.get(canon(sp["display_name"]))
            if rec:
                variant_sprite = {
                    "normal": rec.get("normal"),
                    "shiny":  rec.get("shiny"),
                    "variant_name": rec.get("variant_display"),
                }
        # Manual sprite fallback: if a PNG keyed to this species' slug exists
        # in the variants dir, use it. This catches hand-dropped custom-form
        # art (e.g. Gorochu, Golem) that isn't in the Etrian Variants sheet.
        is_custom_form = False
        if not variant_sprite:
            display_for_slug = DEX_NAME_FIXES.get(canon(dex["name"]), dex["name"]) if dex else sp["display_name"]
            slug_for_file = slugify(display_for_slug)
            normal_file = VARIANTS_DIR / f"{slug_for_file}.png"
            if normal_file.exists():
                shiny_file = VARIANTS_DIR / f"{slug_for_file}-shiny.png"
                variant_sprite = {
                    "normal": f"assets/variants/{slug_for_file}.png",
                    "shiny":  f"assets/variants/{slug_for_file}-shiny.png" if shiny_file.exists() else None,
                    "variant_name": None,
                }
                # Species whose art was manually dropped into the variants
                # directory are custom additions to the game. Tag them so the
                # UI groups them alongside Etrian Variants and Battle Bond forms
                # when filtering.
                is_custom_form = True

        # Apply Pokédex sheet name fixes (typos in the source) and Showdown
        # sprite-slug overrides where our derived slug doesn't match.
        dex_name_raw = dex["name"] if dex else sp["display_name"]
        dex_canon = canon(dex_name_raw)
        display_name = DEX_NAME_FIXES.get(dex_canon, dex_name_raw if dex else sp["display_name"].replace("⭐","").strip().title())
        sprite_slug_value = SPRITE_SLUG_FIXES.get(key) or SPRITE_SLUG_FIXES.get(dex_canon) or sprite_slug(display_name)

        # Linked moves: each gets its slug and feeds the move's used_by[].
        species_slug = slugify(sp["display_name"])
        moves_linked = []
        seen_in_used_by = set()
        for m in sp.get("moves", []):
            nm = (m.get("name") or "").strip()
            docs_slug = slugify(nm) if nm else None
            real_slug = resolve_move_slug(docs_slug) if docs_slug else None
            moves_linked.append({
                "level": m.get("level"),
                "name":  nm,
                "slug":  real_slug,
            })
            if real_slug and (key, real_slug) not in seen_in_used_by:
                seen_in_used_by.add((key, real_slug))
                move_index[real_slug]["used_by"].append({
                    "key":   key,
                    "slug":  species_slug,
                    "name":  display_name,
                    "dex":   dex["dex"] if dex else None,
                    "level": m.get("level"),
                })

        # Linked abilities.
        abilities_linked = []
        for a in sp.get("abilities", []):
            nm = (a or "").strip()
            if not nm:
                continue
            docs_slug = slugify(nm)
            real_slug = resolve_ability_slug(docs_slug)
            abilities_linked.append({"name": nm, "slug": real_slug})
            if real_slug and not any(u["key"] == key for u in ability_index[real_slug]["used_by"]):
                ability_index[real_slug]["used_by"].append({
                    "key":  key,
                    "slug": species_slug,
                    "name": display_name,
                    "dex":  dex["dex"] if dex else None,
                })

        entry = {
            "key": key,
            "display_name": sp["display_name"],
            "name": display_name,
            "slug": species_slug,
            "sprite_slug": sprite_slug_value,
            "dex": dex["dex"] if dex else None,
            "is_variant": is_variant,
            "is_battle_bond": is_battle_bond,
            "is_custom_form": is_custom_form,
            "source_sheet": sp["source_sheet"],
            "types": sp.get("types", []),
            "abilities": abilities_linked,
            "evolves_at": ev,
            "evolves_at_level": ev_level,
            "evolves_note": ev_note,
            "evolution_items": evolution_items,
            "evolution_targets": targets,
            "family": family,
            "stats": stats.get("odyssey") if stats else None,
            "stats_vanilla": stats.get("vanilla") if stats else None,
            "moves": moves_linked,
            "locations": locs,
            "is_event": is_event,
            "has_wild": has_wild,
            "variant_sprite": variant_sprite,
        }
        merged.append(entry)

    def sort_key(e):
        d = e["dex"]
        try: return (0, int(d)) if d else (1, 9999)
        except Exception: return (1, 9999)
    merged.sort(key=sort_key)

    # Write outputs
    items_out = sorted(items.values(), key=lambda x: x["name"].lower())
    (OUT_DIR / "pokedex.json").write_text(
        json.dumps(merged, indent=2, ensure_ascii=False), encoding="utf-8")
    (OUT_DIR / "items.json").write_text(
        json.dumps({"items": items_out, "tutors": tutors}, indent=2, ensure_ascii=False),
        encoding="utf-8")
    moves_out = sorted(move_index.values(), key=lambda x: x["name"].lower())
    abilities_out = sorted(ability_index.values(), key=lambda x: x["name"].lower())
    (OUT_DIR / "moves.json").write_text(
        json.dumps({"moves": moves_out}, indent=2, ensure_ascii=False),
        encoding="utf-8")
    (OUT_DIR / "abilities.json").write_text(
        json.dumps({"abilities": abilities_out}, indent=2, ensure_ascii=False),
        encoding="utf-8")
    (OUT_DIR / "meta.json").write_text(json.dumps({
        "game": "Pokémon Odyssey",
        "version": "v4.1.1",
        "types": ["Normal","Fighting","Flying","Poison","Ground","Rock","Bug","Ghost",
                  "Steel","Fire","Water","Grass","Electric","Psychic","Ice","Dragon","Dark","Aether"],
        "type_chart": type_chart,
        "counts": {
            "species": len(merged),
            "with_stats": sum(1 for e in merged if e["stats"]),
            "with_locations": sum(1 for e in merged if e["locations"]),
            "variants": sum(1 for e in merged if e["is_variant"]),
            "events": sum(1 for e in merged if e["is_event"]),
            "items": len(items_out),
            "moves": len(moves_out),
            "moves_custom": sum(1 for m in moves_out if m.get("is_custom")),
            "abilities": len(abilities_out),
            "abilities_custom": sum(1 for a in abilities_out if a.get("is_custom")),
        },
    }, indent=2, ensure_ascii=False), encoding="utf-8")

    print(f"\nWrote {OUT_DIR / 'pokedex.json'}")
    print(f"Wrote {OUT_DIR / 'items.json'}")
    print(f"Wrote {OUT_DIR / 'moves.json'}")
    print(f"Wrote {OUT_DIR / 'abilities.json'}")
    print(f"Wrote {OUT_DIR / 'meta.json'}")
    if unmatched_species:
        print(f"\n{len(unmatched_species)} species have no dex entry (likely Paradox/variants):")
        for n in unmatched_species[:20]:
            print(f"  - {n}")

if __name__ == "__main__":
    sys.exit(main() or 0)
